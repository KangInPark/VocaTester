from collections import deque
from pathlib import Path
import sys
from openpyxl import load_workbook


class Memorize():
    def __init__(self, cnt):
        if getattr(sys, 'frozen', False):
            self.p = Path(sys.executable).parent.resolve()/"data"
        else:
            self.p = Path(__file__).parent.resolve()/"data"
        self.cnt = cnt
        self.wb = load_workbook(self.p/'review.xlsx')
        self.ws = self.wb[self.wb.sheetnames[-1]]
        self.memolist = deque()
        self.setMemo()
        
    def setMemo(self):
        chk = 0
        for row in self.ws.iter_rows():
            if row[0].value == str(self.cnt) + "차시":
                chk = 1
                continue
            if chk == 0:
                continue
            self.memolist.append((row[0].value, row[1].value, row[2].value, row[3].value))
        self.wb.close()
    
    def nextMemo(self):
        if not self.memolist:
            return None
        else:
            return self.memolist.popleft()
            
