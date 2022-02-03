from pathlib import Path
from random import shuffle
from PyQt5.QtWidgets import *
from PyQt5 import uic, QtCore, QtWidgets
import sys

from openpyxl import load_workbook

from Daily import Daily
from Memorize import Memorize
from Total import Total

p = Path(__file__).parent.resolve()
main_ui = uic.loadUiType(p/"MainWindow.ui")[0]
test_ui = uic.loadUiType(p/"TestWindow.ui")[0]
dailyop_ui = uic.loadUiType(p/"DailyOption.ui")[0]
totalop_ui = uic.loadUiType(p/"TotalOption.ui")[0]
memo_ui = uic.loadUiType(p/"MemorizeWindow.ui")[0]

class MainWindow(QMainWindow, main_ui):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.btn1.clicked.connect(self.DailyBtn)
        self.btn2.clicked.connect(self.TotalBtn)
        if not (p/'data'/'CumulativeWords.xlsx').is_file() and not (p/'data'/'review.xlsx').is_file():
            self.btn2.setEnabled(False)
    
    def DailyBtn(self):
        cnt = 1
        while True:
            self.dailyOption = DailyOption(cnt, 0, None)
            self.dailyOption.exec_()
            if self.dailyOption.userquit == 1 or self.dailyOption.dailyWindow.end == 1:
                return
            self.memorizeWindow = MemorizeWindow(cnt)
            self.memorizeWindow.exec_()
            if self.memorizeWindow.userquit == 1:
                return
            cnt += 1
        
    def TotalBtn(self):
        cnt = 1
        self.totalOption = TotalOption(cnt)
        self.totalOption.exec_()
        if self.totalOption.userquit == 1 or self.totalOption.dailyOption.dailyWindow.end == 1:
            return
        mode = self.totalOption.mode
        self.memorizeWindow = MemorizeWindow(cnt)
        self.memorizeWindow.exec_()
        if self.memorizeWindow.userquit == 1:
            return
        cnt += 1
        while True:
            self.dailyOption = DailyOption(cnt, mode, None)
            self.dailyOption.exec_()
            if self.dailyOption.userquit == 1 or self.dailyOption.dailyWindow.end == 1:
                return
            self.memorizeWindow = MemorizeWindow(cnt)
            self.memorizeWindow.exec_()
            if self.memorizeWindow.userquit == 1:
                return
            cnt += 1
        

class DailyOption(QDialog, dailyop_ui):
    def __init__(self, cnt, mode, scope):
        super().__init__()
        self.setupUi(self)
        self.cnt = cnt
        self.mode = mode
        self.scope = scope
        self.userquit = 1
        self.show()
        self.fbtn.hide()
        self.btn.clicked.connect(self.Btn)
        if cnt == 1 and mode == 0:
            self.fbtn.clicked.connect(self.Fopen)
            self.fbtn.show()
        else:
            self.btn.setEnabled(True)
    
    def Fopen(self):
        path = QtWidgets.QFileDialog.getOpenFileName(self,'OpenFile', filter="Excel files (*.xlsx)")[0]
        if path == "":
            return
        self.fdir = Path(path).resolve()
        self.label.setText('파일 선택 완료')
        self.btn.setEnabled(True)
    
    def Btn(self):
        plist = []
        if self.chk2.isChecked():
            plist.append(1)
        if self.chk3.isChecked():
            plist.append(2)
        if self.chk5.isChecked():
            plist.append(3)
        self.userquit = 0
        self.close()
        if self.cnt == 1 and self.mode == 0:
            info = [self.fdir]
        elif self.cnt > 1:
            info = [p/"data"/"review.xlsx"]
        elif self.mode == 1:
            info = [p/"data"/"CumulativeWords.xlsx"]
            info.extend(self.scope)
        else:
            info = [p/'data'/'review.xlsx']
            info.extend(self.scope)
        tlimit = [0, self.line1.value(), self.line2.value(), self.line3.value()]
        rnd = self.chk1.isChecked()
        same = self.chk4.isChecked()
        self.dailyWindow = TestWindow(self.mode, rnd, plist, info, self.cnt, same, tlimit)
        self.dailyWindow.exec_()
        if self.dailyWindow.userquit == 1:
            self.userquit = 1

class TestWindow(QDialog, test_ui):
    def __init__(self, mode, rnd, plist, info, cnt, same, tlimit):
        super().__init__()
        self.setupUi(self)
        self.cnt = cnt
        self.userquit = 1
        if mode == 0 or cnt > 1:
            self.agent = Daily(rnd, plist, info, self.cnt, same, mode)
        else:
            self.agent = Total(rnd, plist, info, self.cnt, same, mode)
        self.total = len(self.agent.word)
        self.score = 0
        self.n = 0
        self.data = None
        self.wans = []
        self.curr = 0
        self.end = 0
        self.tlimit = tlimit
        self.timer = QtCore.QTimer()
        self.timer.setInterval(500)
        self.elapsed = 0
        self.tover = 0
        self.timer.timeout.connect(self.Timer)
        self.btn1.clicked.connect(self.submit)
        self.btn2.clicked.connect(self.submit)
        self.btn3.clicked.connect(self.submit)
        self.btn4.clicked.connect(self.submit)
        self.btn5.clicked.connect(self.submit)
        self.btn6.clicked.connect(self.submit)
        self.btn7.clicked.connect(self.submit)
        self.btn7.setText("---- 정답을 모르겠음 ----")
        self.btnHide()
        self.lineEdit.hide()
        self.show()
        self.loadQ()
    
    def submit(self):
        self.timer.stop()
        if self.tover == 1:
            self.wans.append((self.data[0], "---- 시간 초과 ----"))
            self.tover = 0
            QtWidgets.QMessageBox.warning(self, "시간 초과", "문제풀이 시간이 초과되었습니다.")
            if self.n == 1:
                self.lineEdit.hide()
                self.lineEdit.setText("")
            elif self.n == 2:
                self.btnHide()
            elif self.n == 3:
                self.btnHide()          
        elif self.n == 1:
            ans = self.lineEdit.text()
            if ans == "":
                return
            if ans == self.data[0][0]:
                self.score += 1
            else:            
                self.wans.append((self.data[0], ans))
            self.lineEdit.hide()
            self.lineEdit.setText("")
        elif self.n == 2:
            if self.sender().text() == self.data[0][2]:
                self.score += 1
            else:
                self.wans.append((self.data[0], self.sender().text()))
            self.btnHide()
        elif self.n == 3:
            if self.sender().text() == self.data[0][0]:
                self.score += 1
            else:
                self.wans.append((self.data[0], self.sender().text()))
            self.btnHide()
        self.loadQ()
    
    def loadQ(self):
        self.n, self.data = self.agent.nextQ()
        if self.n == 0 and self.data == None:
            ment = f'모든 문제를 풀었습니다.\n총 {self.total}문제 중 {self.score}문제를 맞추셨습니다.'
            if self.score != self.total:
                self.agent.review(self.wans)
                ment += f'\n{len(self.wans)}문제의 오답을 점검합니다.'
            else:
                self.end = 1
                ment += '\n오답이 없으므로 학습을 종료합니다.'
            QtWidgets.QMessageBox.information(self, '문제풀이 종료', ment)
            self.userquit = 0
            self.close()
            return
        elif self.n == -1:
            QtWidgets.QMessageBox.critical(self, "경고", "선택한 유형의 테스트를 진행하기에 저장된 단어 데이터의 수가 부족합니다.\n고르기 유형의 경우 최소 6개 이상의 단어 데이터가 필요합니다.\n테스트를 종료합니다.")
            self.end = 1
            self.userquit = 0
            self.close()
            return
        self.curr += 1
        self.setWindowTitle(f'오늘의 단어 ({self.curr}/{self.total})')
        
        if self.n == 1:
            cls = ""
            if self.data[0][1] != None:
                cls = f'[{self.data[0][1]}]'
            self.label.setText(f'{cls}\n{self.data[0][2]}')
            self.ans = self.data[0][0]
            self.lineEdit.raise_()
            self.lineEdit.show()
        elif self.n == 2:
            self.label.setText(self.data[0][0])
            ans = []
            ans.append(self.data[0][2])
            ans.append(self.data[1])
            ans.append(self.data[2])
            ans.append(self.data[3])
            ans.append(self.data[4])
            ans.append(self.data[5])
            shuffle(ans)
            self.btn1.setText(ans[0])
            self.btn2.setText(ans[1])
            self.btn3.setText(ans[2])
            self.btn4.setText(ans[3])
            self.btn5.setText(ans[4])
            self.btn6.setText(ans[5])
            self.btnShow()
        elif self.n == 3:
            txt = ""
            if self.data[0][1] != None:
                txt += f'[{self.data[0][1]}]\n'
            txt += self.data[0][2]
            self.label.setText(txt)
            ans = []
            ans.append(self.data[0][0])
            ans.append(self.data[1])
            ans.append(self.data[2])
            ans.append(self.data[3])
            ans.append(self.data[4])
            ans.append(self.data[5])
            shuffle(ans)
            self.btn1.setText(ans[0])
            self.btn2.setText(ans[1])
            self.btn3.setText(ans[2])
            self.btn4.setText(ans[3])
            self.btn5.setText(ans[4])
            self.btn6.setText(ans[5])
            self.btnShow()
        self.pb.setValue(0)
        self.pb.setMaximum(self.tlimit[self.n]*1000)
        self.elapsed = 0
        self.tlabel.setText(f'{self.tlimit[self.n]}초')
        self.timer.start()

    def btnShow(self):
        self.btn1.show()
        self.btn2.show()
        self.btn3.show()
        self.btn4.show()
        self.btn5.show()
        self.btn6.show()
        self.btn7.show()
        
    def btnHide(self):
        self.btn1.hide()
        self.btn2.hide()
        self.btn3.hide()
        self.btn4.hide() 
        self.btn5.hide() 
        self.btn6.hide() 
        self.btn7.hide() 
    
    def keyReleaseEvent(self, e):
        if e.key() == QtCore.Qt.Key.Key_Enter or e.key() == QtCore.Qt.Key.Key_Return:
            if self.lineEdit.text() != "":
                self.submit()
    
    def Timer(self):
        self.elapsed += 500
        self.pb.setValue(self.elapsed) 
        limit = self.tlimit[self.n] * 1000
        if self.elapsed % 1000 == 0:
            self.tlabel.setText(f'{int((limit - self.elapsed)/1000)}초')
        if limit == self.elapsed:
            self.tover = 1
            self.submit()
                       
class MemorizeWindow(QDialog, memo_ui):
    def __init__(self, cnt):
        super().__init__()
        self.setupUi(self)
        self.cnt = cnt
        self.agent = Memorize(self.cnt)
        self.btn.clicked.connect(self.btnclick)
        self.total = len(self.agent.memolist)
        self.curr = 0
        self.userquit = 1
        self.show()
        self.load()
        
    def load(self):
        self.data = self.agent.nextMemo()
        if self.data == None:
            QtWidgets.QMessageBox.information(self,'오답 점검 완료', '모든 오답을 점검하였습니다. 오답 문제들로 테스트를 재진행합니다.')
            self.userquit = 0
            self.close()
            return
        s = self.data[0] + '\n\n'
        if self.data[1] != None:
            s += f'[{self.data[1]}] '
        s += self.data[2]
        self.label1.setText(s)
        self.label2.setText(self.data[3])
        self.curr += 1
        self.setWindowTitle(f'오답 점검 ({self.curr}/{self.total})')

    def btnclick(self):
        self.load()
    
class TotalOption(QDialog, totalop_ui):
    def __init__(self, cnt):
        super().__init__()
        self.setupUi(self)
        self.cnt = cnt
        self.userquit = 1
        self.show()
        if (p/'data'/'CumulativeWords.xlsx').is_file():
            wb = load_workbook(p/'data'/'CumulativeWords.xlsx')
            self.csheetname = wb.sheetnames
            wb.close()
        else:
            self.rb1.setEnabled(False)
            self.rb2.setChecked(True)
        if (p/'data'/'review.xlsx').is_file():
            wb = load_workbook(p/'data'/'review.xlsx')
            self.rsheetname = wb.sheetnames
            wb.close()
            self.dlist = []
            for sname in self.rsheetname:
                tmp = sname.split('_')[0]
                if tmp not in self.dlist:
                    self.dlist.append(tmp)
        else:
            self.rb2.setEnabled(False)
            self.rb1.setChecked(True)
        self.btn.clicked.connect(self.Btn)
        self.rb1.clicked.connect(self.SetRange1)
        self.rb2.clicked.connect(self.SetRange1)
        self.cb1.activated.connect(self.SetRange2)
        self.SetRange1()
            
    def SetRange1(self):
        self.cb1.clear()
        self.cb2.clear()
        if self.rb1.isChecked():
            self.cb1.addItems(self.csheetname)
        else:
            self.cb1.addItems(self.dlist)
        self.SetRange2()
                   
    def SetRange2(self):
        self.cb2.clear()
        if self.rb1.isChecked():
            tmp = self.csheetname.index(self.cb1.currentText())
            self.cb2.addItems(self.csheetname[tmp:])
        else:
            tmp = self.dlist.index(self.cb1.currentText())
            self.cb2.addItems(self.dlist[tmp:])

    def Btn(self):
        if self.rb1.isChecked():
            self.mode = 1
        else:
            self.mode = 2
        scope = [self.cb1.currentText(), self.cb2.currentText(), self.spinBox.value()]
        self.userquit = 0
        self.close()
        self.dailyOption = DailyOption(self.cnt, self.mode, scope)
        self.dailyOption.exec_()
        if self.dailyOption.userquit == 1:
            self.userquit = 1

if __name__ == '__main__':
    if getattr(sys, 'frozen', False):
        p = Path(sys.executable).parent.resolve()
    q = p/"data"
    if not q.is_dir():
        q.mkdir()
    app = QApplication(sys.argv)
    mainWindow = MainWindow()
    mainWindow.show()
    app.exec_()