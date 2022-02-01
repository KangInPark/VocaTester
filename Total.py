from copy import deepcopy
import datetime
from json import load
import os
import pickle
from random import choice, randint, sample
import sys
from openpyxl import Workbook, load_workbook
from Daily import Daily

class Total(Daily):
    def __init__(self, rnd, plist, info, cnt, same, mode):
        self.rnd = rnd
        self.plist = plist
        self.fdir = info[0]
        self.start = info[1]
        self.end = info[2]
        self.qnum = info[3]
        self.cnt = cnt
        self.same = same
        self.mode = mode
        tmp = load_workbook(os.getcwd() + f'\\data\\CumulativeWords.xlsx')
        pklday = tmp.sheetnames[-1]
        tmp.close()
        self.wb = load_workbook(self.fdir)
        self.setWord()
        self.warn = 0
        if 2 in plist or 3 in plist:
            with open(os.getcwd() + f'\\data\\{pklday}.pkl', 'rb') as f:
                self.mlist = pickle.load(f)
            if len(self.mlist) < 6:
                self.word = []
                self.warn = 1
    
    def setWord(self):
        self.wlist = []
        sname = self.wb.sheetnames
        if self.mode == 2:
            self.start += '_1'
            for s in sname[::-1]:
                if s.split('_')[0] == self.end:
                    self.end += '_' + s.split('_')[1]
                    break
        scope = sname[sname.index(self.start):sname.index(self.end)+1]
        for s in scope:
            self.ws = self.wb[s]
            for row in self.ws.iter_rows():
                if row[0].value == '1차시':
                    continue
                elif row[2].value == None:
                    break
                self.wlist.append((row[0].value, row[1].value, row[2].value))
        if self.qnum > len(self.wlist):
            tmp = deepcopy(self.wlist)
        while self.qnum > len(self.wlist):
            self.wlist.extend(tmp)
        self.word = sample(self.wlist, self.qnum)
        self.wb.close()
