from collections import defaultdict
from copy import deepcopy
from pathlib import Path
import datetime
import pickle
from random import choice, randint
import sys
from openpyxl import Workbook, load_workbook

class Daily():
    def __init__(self, rnd, plist, info, cnt, same, mode):
        if getattr(sys, 'frozen', False):
            self.p = Path(sys.executable).parent.resolve()/"data"
        else:
            self.p = Path(__file__).parent.resolve()/"data"
        self.rnd = rnd
        self.plist = plist
        self.fdir = info[0]
        self.cnt = cnt
        self.same = same
        self.mode = mode
        self.cls = []
        if self.cnt == 1:
            self.wb = load_workbook(self.fdir)
            self.ws = self.wb.active
        else:
            self.wb = load_workbook(self.fdir)
            self.ws = self.wb[self.wb.sheetnames[-1]]
        self.setWord()
        self.warn = 0
        if 2 in plist or 3 in plist:
            pkl = "Wdata.pkl"
            with (self.p/pkl).open('rb') as f:
                self.mlist = list(pickle.load(f))
            if len(self.mlist) < 6:
                self.word = []
                self.warn = 1
            if self.same:
                tmp = defaultdict(int)
                for w in self.mlist:
                    if w[1] in self.cls:
                        tmp[w[1]] += 1
                        if tmp[w[1]] == 6:
                            self.cls.remove(w[1])
                            if not self.cls:
                                break
                 
    def setWord(self):
        self.word = []
        chk = 1
        if self.cnt >= 2:
            chk = 0
        for row in self.ws.iter_rows():
            if row[0].value == str(self.cnt-1) + "차시":
                chk = 1
                continue
            if chk == 0:
                continue
            self.word.append((row[0].value, row[1].value, row[2].value))
            if(row[1].value != None and row[1].value not in self.cls):
                self.cls.append(row[1].value)
        if self.mode == 0:
            self.accumulate(self.word)
        self.wb.close()
    
    def nextQ(self):
        if not self.word:
            if self.warn == 1:
                return -1, None
            return 0, None
        if self.rnd:
            idx = randint(0,len(self.word)-1)
        else:
            idx = 0
        n = choice(self.plist)
        tmp = self.word.pop(idx)
        ret = []
        if n == 1:
            ret.append(tmp)
        elif n == 2:
            ret.append(tmp)
            while len(ret) < 6:
                pick = choice(self.mlist)
                sam = pick[2]
                if sam != ret[0][2] and sam not in ret:
                    if self.same and ret[0][1] not in self.cls:
                        if ret[0][1] == None:
                            ret.append(sam)
                        else:
                            if ret[0][1] == pick[1]:
                                ret.append(sam)
                    else:
                        ret.append(sam)
        elif n == 3:
            ret.append(tmp)
            while len(ret) < 6:
                pick = choice(self.mlist)
                sam = pick[0]
                if sam != ret[0][0] and sam not in ret:
                    if self.same and ret[0][1] not in self.cls:
                        if ret[0][1] == None:
                            ret.append(sam)
                        else:
                            if ret[0][1] == pick[1]:
                                ret.append(sam)
                    else:
                        ret.append(sam)
        return n, ret
    
    def review(self, wans):
        path = self.p/'review.xlsx'
        if not path.is_file():
            wb = Workbook()
            ws = wb['Sheet']
            ws.title = f'{str(datetime.date.today())}_1'
        else:
            wb = load_workbook(path)
            chk = wb.sheetnames[-1].split('_')
            if  chk[0] != str(datetime.date.today()):    
                sname = f'{str(datetime.date.today())}_1'
                ws = wb.create_sheet(sname)
            elif self.cnt == 1:
                sname = f'{chk[0]}_{str(int(chk[1])+1)}'
                ws = wb.create_sheet(sname)
            else:
                ws = wb[self.wb.sheetnames[-1]]
        row = 1
        if self.cnt == 1:
            ws.cell(row,1).value = "1차시"
            row += 1
        else:
            while ws.cell(row,1).value != None:
                row += 1
            ws.cell(row,1).value = str(self.cnt)+"차시"
            row += 1            
        for i in range(row, len(wans) + row):
            ws.cell(i,1).value = wans[i-row][0][0]
            if wans[i-row][0][1] == None:
                ws.cell(i,2).value = ""
            else:
                ws.cell(i,2).value = wans[i-row][0][1]
            ws.cell(i,3).value = wans[i-row][0][2]
            ws.cell(i,4).value = wans[i-row][1]
        wb.save(path)
        wb.close()
        
    def accumulate(self, words):
        word = deepcopy(words)
        path = self.p/'CumulativeWords.xlsx'
        dup = {}
        today = str(datetime.date.today())
        if not path.is_file():
            wb = Workbook()
            ws = wb['Sheet']
            ws.title = today
        else:
            if not (self.p/'Wdata.pkl').is_file():
                self.makePkl()
            wb = load_workbook(path)
            if today in wb.sheetnames:
                ws = wb[today]
                for item in word:
                    if item[0] not in dup:
                        dup[item[0]] = 1
                for row in ws.iter_rows():
                    if row[0].value not in dup:
                        word.append((row[0].value, row[1].value, row[2].value))
                del wb[today]
            ws = wb.create_sheet(today)
        for i in range(1, len(word)+1):
            ws.cell(i,1).value = word[i-1][0]
            if word[i-1][1] == None:
                ws.cell(i,2).value = ""
            else:
                ws.cell(i,2).value = word[i-1][1]
            ws.cell(i,3).value = word[i-1][2]
        wb.save(path)
        pick = set()
        if (self.p/'Wdata.pkl').is_file():
            with (self.p/'Wdata.pkl').open('rb') as f:
                pick = pickle.load(f)
        pick.update(word)
        with (self.p/'Wdata.pkl').open('wb') as f:
            pickle.dump(pick, f)
        wb.close()
        
    def makePkl(self):
        with (self.p/'Wdata.pkl').open('wb') as f:
            wb = load_workbook(self.p/'CumulativeWords.xlsx')
            tmp = wb.sheetnames
            data = set()
            for sname in tmp:
                ws = wb[sname]
                for row in ws.iter_rows():
                    data.add((row[0].value, row[1].value, row[2].value))
            pickle.dump(data, f)
            wb.close()